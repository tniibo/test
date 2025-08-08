import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill, Border, Side, Alignment
import matplotlib.pyplot as plt
import matplotlib.patches as patches
from matplotlib.patches import Rectangle
import matplotlib.font_manager as fm
import numpy as np
from pathlib import Path
import json
import base64
from io import BytesIO
from PIL import Image
import warnings
warnings.filterwarnings('ignore')

# 日本語フォントの設定
def setup_japanese_font():
    """日本語フォントを設定（Azure Linux環境対応）"""
    import os
    import platform
    
    # 利用可能な日本語フォントを探す（優先順位付き）
    japanese_fonts = [
        'Noto Sans CJK JP',  # Azure/Linux推奨
        'Noto Sans JP',  # Cross-platform
        'IPAexGothic',  # Linux/Unix
        'IPAPGothic',  # Linux/Unix
        'IPAGothic',  # Linux/Unix
        'TakaoGothic',  # Linux/Unix
        'Takao',  # Linux/Unix
        'VL Gothic',  # Linux
        'MS Gothic',  # Windows
        'MS Mincho',  # Windows
        'Yu Gothic',  # Windows
        'Hiragino Sans',  # Mac
        'Hiragino Maru Gothic Pro',  # Mac
        'DejaVu Sans'  # Fallback
    ]
    
    # Azure環境かどうかチェック
    is_azure = os.environ.get('WEBSITE_INSTANCE_ID') is not None or \
               os.environ.get('AZURE_FUNCTIONS_ENVIRONMENT') is not None
    
    # Linux環境での追加フォントパス
    if platform.system() == 'Linux' or is_azure:
        # Azure App Service/Functions用のフォントパス
        additional_font_paths = [
            '/usr/share/fonts/opentype/noto/',
            '/usr/share/fonts/truetype/noto/',
            '/usr/share/fonts/truetype/liberation/',
            '/usr/local/share/fonts/',
            '/home/site/wwwroot/fonts/',  # カスタムフォント用
            '/opt/fonts/',  # カスタムフォント用
        ]
        
        # フォントパスを追加
        for path in additional_font_paths:
            if os.path.exists(path):
                fm.fontManager.addfont(path)
    
    available_fonts = fm.findSystemFonts()
    font_found = False
    
    # フォントパスから直接フォントを探す
    for font_path in available_fonts:
        font_basename = os.path.basename(font_path).lower()
        for font_name in japanese_fonts:
            if font_name.lower().replace(' ', '') in font_basename.replace('-', '').replace('_', ''):
                try:
                    # フォントプロパティを直接設定
                    prop = fm.FontProperties(fname=font_path)
                    plt.rcParams['font.family'] = prop.get_name()
                    font_found = True
                    print(f"使用フォント: {font_name} ({font_path})")
                    break
                except:
                    continue
        if font_found:
            break
    
    # フォント名で検索（fallback）
    if not font_found:
        for font_name in japanese_fonts:
            try:
                plt.rcParams['font.family'] = [font_name]
                # テスト描画で確認
                fig, ax = plt.subplots(figsize=(1, 1))
                ax.text(0.5, 0.5, 'テスト', fontsize=12)
                plt.close(fig)
                font_found = True
                print(f"使用フォント: {font_name}")
                break
            except:
                continue
    
    if not font_found:
        # フォントが見つからない場合の設定
        plt.rcParams['font.family'] = ['sans-serif']
        plt.rcParams['font.sans-serif'] = ['DejaVu Sans']
        
        print("警告: 日本語フォントが見つかりません。文字化けする可能性があります。")
        print("\n=== Azure/Linux環境での日本語フォントインストール方法 ===")
        print("1. Dockerfileを使用する場合:")
        print("   RUN apt-get update && apt-get install -y fonts-noto-cjk fonts-ipafont-gothic")
        print("\n2. Azure App Serviceの場合:")
        print("   - スタートアップコマンドに追加:")
        print("   apt-get update && apt-get install -y fonts-noto-cjk")
        print("\n3. カスタムフォントを使用する場合:")
        print("   - /home/site/wwwroot/fonts/ にフォントファイルを配置")
        print("=" * 60)
    
    # 負の符号の表示設定
    plt.rcParams['axes.unicode_minus'] = False

# グローバルでフォント設定を実行
setup_japanese_font()

class ExcelToMarkdownPreprocessor:
    """ExcelファイルをMarkdown変換用に最適化して抽出するクラス"""
    
    def __init__(self, file_path):
        """
        初期化
        
        Args:
            file_path (str): Excelファイルのパス
        """
        self.file_path = Path(file_path)
        self.workbook = openpyxl.load_workbook(file_path, data_only=True)
        self.workbook_with_formulas = openpyxl.load_workbook(file_path, data_only=False)
        self.sheet_names = self.workbook.sheetnames
        self.ai_optimized_data = {}
        
        # 全シートのデータを処理
        for sheet_name in self.sheet_names:
            self.ai_optimized_data[sheet_name] = self.extract_structured_data(sheet_name)
        
    def detect_table_structure(self, sheet):
        """
        テーブル構造を自動検出
        
        Args:
            sheet: openpyxlのシートオブジェクト
            
        Returns:
            dict: テーブル構造情報
        """
        tables = []
        visited = set()
        
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None and cell.coordinate not in visited:
                    # 連続したデータ領域を検出
                    table = self._find_continuous_region(sheet, cell.row, cell.column, visited)
                    if table['cell_count'] > 1:  # 2セル以上のまとまりをテーブルとして認識
                        tables.append(table)
        
        return tables
    
    def _find_continuous_region(self, sheet, start_row, start_col, visited):
        """
        連続したデータ領域を検出
        
        Args:
            sheet: シートオブジェクト
            start_row: 開始行
            start_col: 開始列
            visited: 訪問済みセルのセット
            
        Returns:
            dict: テーブル領域情報
        """
        min_row, max_row = start_row, start_row
        min_col, max_col = start_col, start_col
        cells_in_region = []
        
        # BFSで連続領域を探索
        queue = [(start_row, start_col)]
        region_visited = set()
        
        while queue:
            row, col = queue.pop(0)
            if (row, col) in region_visited:
                continue
                
            cell = sheet.cell(row=row, column=col)
            if cell.value is not None:
                region_visited.add((row, col))
                visited.add(cell.coordinate)
                cells_in_region.append({
                    'row': row,
                    'col': col,
                    'value': cell.value,
                    'coordinate': cell.coordinate
                })
                
                min_row = min(min_row, row)
                max_row = max(max_row, row)
                min_col = min(min_col, col)
                max_col = max(max_col, col)
                
                # 隣接セルをチェック（空白1セルまでは許容）
                for dr in [-2, -1, 0, 1, 2]:
                    for dc in [-2, -1, 0, 1, 2]:
                        if dr == 0 and dc == 0:
                            continue
                        new_row, new_col = row + dr, col + dc
                        if (new_row > 0 and new_col > 0 and 
                            (new_row, new_col) not in region_visited):
                            queue.append((new_row, new_col))
        
        # ヘッダー行を推定
        potential_header_row = min_row
        has_header = self._detect_header_row(sheet, min_row, min_col, max_col)
        
        return {
            'bounds': {
                'min_row': min_row,
                'max_row': max_row,
                'min_col': min_col,
                'max_col': max_col
            },
            'cell_count': len(cells_in_region),
            'cells': cells_in_region,
            'has_header': has_header,
            'header_row': potential_header_row if has_header else None,
            'type': 'table' if has_header else 'data_region'
        }
    
    def _detect_header_row(self, sheet, row, min_col, max_col):
        """
        ヘッダー行を検出
        
        Args:
            sheet: シートオブジェクト
            row: チェックする行
            min_col: 開始列
            max_col: 終了列
            
        Returns:
            bool: ヘッダー行かどうか
        """
        # ヘッダー行の特徴：文字列が多い、太字、背景色など
        text_count = 0
        total_count = 0
        
        for col in range(min_col, max_col + 1):
            cell = sheet.cell(row=row, column=col)
            if cell.value is not None:
                total_count += 1
                if isinstance(cell.value, str):
                    text_count += 1
                    
        # 70%以上が文字列ならヘッダーと判定
        return text_count / total_count > 0.7 if total_count > 0 else False
    
    def extract_structured_data(self, sheet_name=None, use_print_area=True):
        """
        AI用に最小限の構造化データを抽出（印刷範囲対応）
        
        Args:
            sheet_name (str): シート名（Noneの場合は全シート）
            use_print_area (bool): 印刷範囲のみを対象とするか
            
        Returns:
            dict: 最小化されたデータ
        """
        if sheet_name:
            sheets_to_process = [sheet_name] if sheet_name in self.sheet_names else []
        else:
            sheets_to_process = self.sheet_names
            
        result_data = {}
            
        for sheet_name in sheets_to_process:
            sheet = self.workbook[sheet_name]
            
            # 印刷範囲を取得
            print_area = None
            print_area_bounds = None
            if use_print_area and sheet.print_area:
                print_area = sheet.print_area
                # 印刷範囲をパース（例: "A1:Z100" or "Sheet1!$A$1:$Z$100"）
                if '!' in print_area:
                    print_area = print_area.split('!')[-1]
                print_area = print_area.replace('$', '')
                
                # 範囲を解析
                if ':' in print_area:
                    import re
                    range_parts = print_area.split(':')
                    start_match = re.match(r'([A-Z]+)(\d+)', range_parts[0])
                    end_match = re.match(r'([A-Z]+)(\d+)', range_parts[1])
                    if start_match and end_match:
                        from openpyxl.utils import column_index_from_string
                        print_area_bounds = {
                            'min_row': int(start_match.group(2)),
                            'max_row': int(end_match.group(2)),
                            'min_col': column_index_from_string(start_match.group(1)),
                            'max_col': column_index_from_string(end_match.group(1))
                        }
            
            # 最小限のシートデータ構造
            sheet_data = {
                'cells': {},  # {座標: 値} の辞書
                'merged': [],  # 結合セルのリスト
                'tables': [],  # テーブル領域のリスト
                'print_area': print_area,  # 印刷範囲
                'data_bounds': None,  # 実データ範囲
                '_internal': {  # 内部処理用
                    'tables_detail': [],
                    'formatting_hints': [],
                    'semantic_structure': {'document_type': 'unknown', 'sections': []},
                    'print_area_bounds': print_area_bounds
                }
            }
            
            # データ範囲を記録
            min_row, max_row = float('inf'), 0
            min_col, max_col = float('inf'), 0
            
            # すべての値があるセルを記録（印刷範囲内のみ）
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        # 印刷範囲チェック
                        if print_area_bounds:
                            if not (print_area_bounds['min_row'] <= cell.row <= print_area_bounds['max_row'] and
                                   print_area_bounds['min_col'] <= cell.column <= print_area_bounds['max_col']):
                                continue
                        
                        # 座標をキー、値を値とするシンプルな辞書
                        sheet_data['cells'][cell.coordinate] = cell.value
                        
                        # データ範囲を更新
                        min_row = min(min_row, cell.row)
                        max_row = max(max_row, cell.row)
                        min_col = min(min_col, cell.column)
                        max_col = max(max_col, cell.column)
            
            # データ範囲を記録
            if sheet_data['cells']:
                from openpyxl.utils import get_column_letter
                sheet_data['data_bounds'] = {
                    'min_row': min_row,
                    'max_row': max_row,
                    'min_col': min_col,
                    'max_col': max_col,
                    'range': f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
                }
            
            result_data[sheet_name] = sheet_data
            
        # 単一シートの場合はそのデータを、複数シートの場合は全体を返す
        if len(result_data) == 1:
            return list(result_data.values())[0]
        else:
            return result_data
    
    def _extract_table_data(self, sheet, table_info):
        """
        テーブルデータを構造化して抽出（空セルは除外）
        
        Args:
            sheet: シートオブジェクト
            table_info: テーブル情報
            
        Returns:
            dict: 構造化されたテーブルデータ
        """
        bounds = table_info['bounds']
        table_data = {
            'bounds': bounds,
            'has_header': table_info['has_header'],
            'data': [],
            'headers': [],
            'data_types': {},
            'alignment': 'unknown',
            'non_empty_cells_count': 0
        }
        
        # ヘッダーの抽出
        if table_info['has_header']:
            for col in range(bounds['min_col'], bounds['max_col'] + 1):
                cell = sheet.cell(row=bounds['min_row'], column=col)
                if cell.value is not None:  # 空でないヘッダーのみ
                    table_data['headers'].append({
                        'value': str(cell.value),
                        'column': get_column_letter(col),
                        'column_index': col
                    })
        
        # データ行の抽出
        start_row = bounds['min_row'] + 1 if table_info['has_header'] else bounds['min_row']
        for row_idx in range(start_row, bounds['max_row'] + 1):
            row_data = []
            has_data = False  # この行にデータがあるかチェック
            
            for col in range(bounds['min_col'], bounds['max_col'] + 1):
                cell = sheet.cell(row=row_idx, column=col)
                
                # 値がある場合のみデータを追加
                if cell.value is not None:
                    cell_data = {
                        'row': row_idx,
                        'column': col,
                        'value': cell.value,
                        'type': type(cell.value).__name__,
                        'formatted_value': self._format_cell_value(cell),
                        'coordinate': cell.coordinate
                    }
                    
                    # データ型の統計
                    dtype = cell_data['type']
                    table_data['data_types'][dtype] = table_data['data_types'].get(dtype, 0) + 1
                    table_data['non_empty_cells_count'] += 1
                    
                    row_data.append(cell_data)
                    has_data = True
            
            # データがある行のみ追加
            if has_data:
                table_data['data'].append({
                    'row_index': row_idx,
                    'cells': row_data
                })
        
        # テーブルの向きを推定（縦持ち or 横持ち）
        if len(table_data['headers']) > 0:
            if len(table_data['data']) > len(table_data['headers']):
                table_data['alignment'] = 'vertical'
            else:
                table_data['alignment'] = 'horizontal'
        
        return table_data
    
    def _format_cell_value(self, cell):
        """
        セルの値をMarkdown用にフォーマット
        
        Args:
            cell: セルオブジェクト
            
        Returns:
            str: フォーマットされた値
        """
        if cell.value is None:
            return ''
        elif isinstance(cell.value, bool):
            return 'Yes' if cell.value else 'No'
        elif isinstance(cell.value, (int, float)):
            if cell.number_format and '%' in str(cell.number_format):
                return f"{cell.value * 100:.1f}%"
            elif isinstance(cell.value, float):
                return f"{cell.value:.2f}"
            else:
                return str(cell.value)
        else:
            return str(cell.value)
    
    def _extract_formatting_hints(self, sheet):
        """
        書式設定からMarkdown変換のヒントを抽出
        
        Args:
            sheet: シートオブジェクト
            
        Returns:
            list: フォーマットヒント
        """
        hints = []
        
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    hint = {}
                    
                    # 太字・斜体
                    if cell.font:
                        if cell.font.bold:
                            hint['bold'] = True
                        if cell.font.italic:
                            hint['italic'] = True
                        if cell.font.size and cell.font.size > 12:
                            hint['large_text'] = True
                            hint['possible_heading'] = True
                    
                    # 背景色
                    if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                        hint['has_background'] = True
                        hint['emphasis'] = True
                    
                    # 配置
                    if cell.alignment:
                        if cell.alignment.horizontal == 'center':
                            hint['centered'] = True
                        if cell.alignment.vertical == 'top':
                            hint['top_aligned'] = True
                    
                    if hint:
                        hint['coordinate'] = cell.coordinate
                        hint['value'] = cell.value
                        hints.append(hint)
        
        return hints
    
    def _infer_semantic_structure(self, sheet_data):
        """
        データからセマンティック構造を推定
        
        Args:
            sheet_data: シートデータ
            
        Returns:
            dict: セマンティック構造
        """
        structure = {
            'document_type': 'unknown',
            'sections': [],
            'key_value_pairs': [],
            'lists': []
        }
        
        # テーブルの数と種類から文書タイプを推定
        table_count = len(sheet_data['tables'])
        if table_count == 0:
            structure['document_type'] = 'text_document'
        elif table_count == 1:
            structure['document_type'] = 'single_table'
        else:
            structure['document_type'] = 'multi_table_report'
        
        # セクションの検出（結合セルがタイトルの可能性）
        for merged in sheet_data['merged_regions']:
            if merged['value'] and isinstance(merged['value'], str):
                structure['sections'].append({
                    'title': merged['value'],
                    'level': 1 if merged['span']['cols'] > 3 else 2,
                    'position': merged['bounds']
                })
        
        # キー・バリューペアの検出
        for table in sheet_data['tables']:
            if table['alignment'] == 'horizontal' and len(table['data']) <= 3:
                # 横持ちで行数が少ない場合はキー・バリューの可能性
                structure['key_value_pairs'].append({
                    'table_id': table.get('table_id'),
                    'type': 'properties'
                })
        
        return structure
    
    def create_paged_visualizations_with_data(self, sheet_name, output_dir="ai_input", 
                                             rows_per_page=30, cols_per_page=10, 
                                             overlap=2, dpi=120):
        """
        シートをページ単位に分割して可視化（データも紐付け）
        
        Args:
            sheet_name (str): シート名
            output_dir (str): 出力ディレクトリ
            rows_per_page (int): 1ページあたりの行数
            cols_per_page (int): 1ページあたりの列数
            overlap (int): ページ間のオーバーラップ
            dpi (int): 画像の解像度
            
        Returns:
            dict: ページごとのデータと画像パス
        """
        from pathlib import Path
        output_path = Path(output_dir)
        output_path.mkdir(exist_ok=True)
        
        if sheet_name not in self.ai_optimized_data:
            self.extract_structured_data(sheet_name)
        
        sheet = self.workbook[sheet_name]
        sheet_data = self.ai_optimized_data[sheet_name]
        
        # sheet_dataがNoneの場合の処理
        if sheet_data is None:
            print(f"警告: シート '{sheet_name}' のデータが処理されていません。")
            return []
        
        # データ範囲または印刷範囲を使用
        if sheet_data.get('data_bounds'):
            min_data_row = sheet_data['data_bounds']['min_row']
            max_data_row = sheet_data['data_bounds']['max_row']
            min_data_col = sheet_data['data_bounds']['min_col']
            max_data_col = sheet_data['data_bounds']['max_col']
        else:
            return {}  # データがない場合
        
        # ページ分割の計算
        num_row_pages = ((max_data_row - min_data_row + 1) + rows_per_page - 1) // rows_per_page
        num_col_pages = ((max_data_col - min_data_col + 1) + cols_per_page - 1) // cols_per_page
        
        pages_data = {}
        page_index = 0
        
        # 各ページを生成
        for row_page in range(num_row_pages):
            for col_page in range(num_col_pages):
                page_index += 1
                
                # ページの範囲を計算
                page_min_row = min_data_row + row_page * rows_per_page
                page_max_row = min(page_min_row + rows_per_page - 1, max_data_row)
                page_min_col = min_data_col + col_page * cols_per_page
                page_max_col = min(page_min_col + cols_per_page - 1, max_data_col)
                
                # オーバーラップを追加（次のページとの連続性のため）
                if row_page < num_row_pages - 1:
                    page_max_row = min(page_max_row + overlap, max_data_row)
                if col_page < num_col_pages - 1:
                    page_max_col = min(page_max_col + overlap, max_data_col)
                
                # ページのセルデータを抽出
                page_cells = {}
                page_tables = []
                page_merged = []
                
                # このページに含まれるセルを抽出
                for coord, value in sheet_data['cells'].items():
                    import re
                    match = re.match(r'([A-Z]+)(\d+)', coord)
                    if match:
                        col = column_index_from_string(match.group(1))
                        row = int(match.group(2))
                        if page_min_row <= row <= page_max_row and page_min_col <= col <= page_max_col:
                            page_cells[coord] = value
                
                # このページに含まれる/交差するテーブルを特定
                for table in sheet_data['tables']:
                    import re
                    range_parts = table['range'].split(':')
                    if len(range_parts) == 2:
                        start_match = re.match(r'([A-Z]+)(\d+)', range_parts[0])
                        end_match = re.match(r'([A-Z]+)(\d+)', range_parts[1])
                        if start_match and end_match:
                            t_min_row = int(start_match.group(2))
                            t_max_row = int(end_match.group(2))
                            t_min_col = column_index_from_string(start_match.group(1))
                            t_max_col = column_index_from_string(end_match.group(1))
                            
                            # テーブルがページと交差するかチェック
                            if (t_min_row <= page_max_row and t_max_row >= page_min_row and
                                t_min_col <= page_max_col and t_max_col >= page_min_col):
                                page_tables.append(table)
                
                # このページに含まれる結合セルを特定
                for merged in sheet_data['merged']:
                    import re
                    range_match = re.match(r'([A-Z]+)(\d+):([A-Z]+)(\d+)', merged['range'])
                    if range_match:
                        m_min_col = column_index_from_string(range_match.group(1))
                        m_min_row = int(range_match.group(2))
                        m_max_col = column_index_from_string(range_match.group(3))
                        m_max_row = int(range_match.group(4))
                        
                        if (m_min_row <= page_max_row and m_max_row >= page_min_row and
                            m_min_col <= page_max_col and m_max_col >= page_min_col):
                            page_merged.append(merged)
                
                # ページ画像を生成
                page_file = output_path / f"{sheet_name}_page_{page_index:03d}.png"
                
                fig = self._create_page_visualization(
                    sheet, sheet_data,
                    page_min_row, page_max_row,
                    page_min_col, page_max_col,
                    page_index, num_row_pages * num_col_pages,
                    dpi
                )
                
                if fig:
                    plt.figure(fig.number)
                    plt.savefig(page_file, dpi=dpi, bbox_inches='tight',
                               facecolor='white', edgecolor='none')
                    plt.close(fig)
                
                # ページデータを記録
                page_key = f"page_{page_index:03d}"
                pages_data[page_key] = {
                    'image': page_file.name,
                    'range': f"{get_column_letter(page_min_col)}{page_min_row}:{get_column_letter(page_max_col)}{page_max_row}",
                    'cells': page_cells,
                    'tables': page_tables,
                    'merged': page_merged,
                    'stats': {
                        'cell_count': len(page_cells),
                        'table_count': len(page_tables),
                        'merged_count': len(page_merged)
                    }
                }
                
                print(f"  ページ {page_index}/{num_row_pages * num_col_pages}: {page_file.name} (セル数: {len(page_cells)})")
        
        return pages_data
        """
        シートをページ単位に分割して可視化
        
        Args:
            sheet_name (str): シート名
            output_dir (str): 出力ディレクトリ
            rows_per_page (int): 1ページあたりの行数
            cols_per_page (int): 1ページあたりの列数
            overlap (int): ページ間のオーバーラップ（連続性確保用）
            dpi (int): 画像の解像度
            
        Returns:
            list: 生成された画像ファイルパスのリスト
        """
        from pathlib import Path
        output_path = Path(output_dir)
        output_path.mkdir(exist_ok=True)
        
        if sheet_name not in self.ai_optimized_data:
            self.extract_structured_data(sheet_name)
        
        sheet = self.workbook[sheet_name]
        sheet_data = self.ai_optimized_data[sheet_name]
        
        # データがある範囲を検出
        min_data_row = float('inf')
        max_data_row = 0
        min_data_col = float('inf')
        max_data_col = 0
        
        for coord in sheet_data['cells'].keys():
            import re
            match = re.match(r'([A-Z]+)(\d+)', coord)
            if match:
                col = column_index_from_string(match.group(1))
                row = int(match.group(2))
                min_data_row = min(min_data_row, row)
                max_data_row = max(max_data_row, row)
                min_data_col = min(min_data_col, col)
                max_data_col = max(max_data_col, col)
        
        # ページ分割の計算
        if min_data_row == float('inf'):  # データがない場合
            return []
        
        # ページ数を計算
        num_row_pages = ((max_data_row - min_data_row + 1) + rows_per_page - 1) // rows_per_page
        num_col_pages = ((max_data_col - min_data_col + 1) + cols_per_page - 1) // cols_per_page
        
        generated_files = []
        page_index = 0
        
        # 各ページを生成
        for row_page in range(num_row_pages):
            for col_page in range(num_col_pages):
                page_index += 1
                
                # ページの範囲を計算（オーバーラップを考慮）
                page_min_row = min_data_row + row_page * rows_per_page
                page_max_row = min(page_min_row + rows_per_page - 1 + overlap, max_data_row)
                page_min_col = min_data_col + col_page * cols_per_page
                page_max_col = min(page_min_col + cols_per_page - 1 + overlap, max_data_col)
                
                # ページ画像を生成
                page_file = output_path / f"{sheet_name}_page_{page_index:03d}.png"
                
                fig = self._create_page_visualization(
                    sheet, sheet_data,
                    page_min_row, page_max_row,
                    page_min_col, page_max_col,
                    page_index, num_row_pages * num_col_pages,
                    dpi
                )
                
                if fig:
                    plt.figure(fig.number)
                    plt.savefig(page_file, dpi=dpi, bbox_inches='tight',
                               facecolor='white', edgecolor='none')
                    plt.close(fig)
                    generated_files.append(str(page_file))
                    print(f"  ページ {page_index}/{num_row_pages * num_col_pages}: {page_file.name}")
        
        return generated_files
    
    def _create_page_visualization(self, sheet, sheet_data, 
                                 page_min_row, page_max_row,
                                 page_min_col, page_max_col,
                                 page_num, total_pages, dpi):
        """
        単一ページの可視化を作成
        
        Args:
            sheet: シートオブジェクト
            sheet_data: シートデータ
            page_min_row: ページ最小行
            page_max_row: ページ最大行
            page_min_col: ページ最小列
            page_max_col: ページ最大列
            page_num: ページ番号
            total_pages: 総ページ数
            dpi: 解像度
            
        Returns:
            matplotlib.figure.Figure: 生成された図
        """
        num_rows = page_max_row - page_min_row + 1
        num_cols = page_max_col - page_min_col + 1
        
        # セルサイズ
        cell_width = 2.5
        cell_height = 0.8
        
        # 図のサイズを計算（A4サイズを意識）
        fig_width = min(num_cols * cell_width, 20)  # 最大幅20インチ
        fig_height = min(num_rows * cell_height, 14)  # 最大高さ14インチ
        
        fig, ax = plt.subplots(figsize=(fig_width, fig_height))
        
        # 座標系の設定
        ax.set_xlim(0, num_cols * cell_width)
        ax.set_ylim(0, num_rows * cell_height)
        
        # 背景グリッド
        for i in range(num_rows + 1):
            ax.axhline(y=i * cell_height, color='#D0D0D0', linewidth=0.8)
        for j in range(num_cols + 1):
            ax.axvline(x=j * cell_width, color='#D0D0D0', linewidth=0.8)
        
        # テーブル領域を描画
        colors = ['#E8F4FD', '#E8F5E9', '#FFF8E1', '#FCE4EC', '#F3E5F5']
        
        # テーブル詳細データを取得
        if '_internal' in sheet_data and 'tables_detail' in sheet_data['_internal']:
            tables_detail = sheet_data['_internal']['tables_detail']
        else:
            tables_detail = []
            for table_info in sheet_data['tables']:
                range_parts = table_info['range'].split(':')
                if len(range_parts) == 2:
                    import re
                    start_match = re.match(r'([A-Z]+)(\d+)', range_parts[0])
                    end_match = re.match(r'([A-Z]+)(\d+)', range_parts[1])
                    if start_match and end_match:
                        tables_detail.append({
                            'bounds': {
                                'min_row': int(start_match.group(2)),
                                'max_row': int(end_match.group(2)),
                                'min_col': column_index_from_string(start_match.group(1)),
                                'max_col': column_index_from_string(end_match.group(1))
                            }
                        })
        
        # ページ内のテーブルを描画
        for idx, table in enumerate(tables_detail):
            if 'bounds' in table:
                bounds = table['bounds']
                # ページ範囲と交差するテーブルのみ描画
                if (bounds['min_row'] <= page_max_row and 
                    bounds['max_row'] >= page_min_row and
                    bounds['min_col'] <= page_max_col and
                    bounds['max_col'] >= page_min_col):
                    
                    # ページ座標に変換
                    table_min_row = max(bounds['min_row'], page_min_row) - page_min_row
                    table_max_row = min(bounds['max_row'], page_max_row) - page_min_row
                    table_min_col = max(bounds['min_col'], page_min_col) - page_min_col
                    table_max_col = min(bounds['max_col'], page_max_col) - page_min_col
                    
                    rect = Rectangle(
                        (table_min_col * cell_width,
                         (num_rows - table_max_row - 1) * cell_height),
                        (table_max_col - table_min_col + 1) * cell_width,
                        (table_max_row - table_min_row + 1) * cell_height,
                        facecolor=colors[idx % len(colors)],
                        edgecolor='#2196F3',
                        linewidth=2,
                        alpha=0.3
                    )
                    ax.add_patch(rect)
                    
                    # テーブルラベル（ページ内の左上に）
                    ax.text(
                        table_min_col * cell_width + 0.1,
                        (num_rows - table_min_row) * cell_height - 0.1,
                        f"Table {idx+1}",
                        fontsize=10,
                        fontweight='bold',
                        color='#1565C0',
                        bbox=dict(boxstyle="round,pad=0.2",
                                facecolor='white',
                                edgecolor='#1565C0',
                                alpha=0.9)
                    )
        
        # 結合セルを描画
        for merged in sheet_data['merged']:
            import re
            range_match = re.match(r'([A-Z]+)(\d+):([A-Z]+)(\d+)', merged['range'])
            if range_match:
                m_min_col = column_index_from_string(range_match.group(1))
                m_min_row = int(range_match.group(2))
                m_max_col = column_index_from_string(range_match.group(3))
                m_max_row = int(range_match.group(4))
                
                # ページ範囲内の場合のみ描画
                if (m_min_row <= page_max_row and 
                    m_max_row >= page_min_row and
                    m_min_col <= page_max_col and
                    m_max_col >= page_min_col):
                    
                    # ページ座標に変換
                    merge_min_row = max(m_min_row, page_min_row) - page_min_row
                    merge_max_row = min(m_max_row, page_max_row) - page_min_row
                    merge_min_col = max(m_min_col, page_min_col) - page_min_col
                    merge_max_col = min(m_max_col, page_max_col) - page_min_col
                    
                    rect = Rectangle(
                        (merge_min_col * cell_width,
                         (num_rows - merge_max_row - 1) * cell_height),
                        (merge_max_col - merge_min_col + 1) * cell_width,
                        (merge_max_row - merge_min_row + 1) * cell_height,
                        facecolor='#FFF59D',
                        edgecolor='#F57C00',
                        linewidth=1.5,
                        alpha=0.4
                    )
                    ax.add_patch(rect)
                    
                    # 結合セルの値を表示
                    if merged['value']:
                        text = str(merged['value'])
                        if len(text) > 50:
                            text = text[:47] + '...'
                        
                        center_col = (merge_min_col + merge_max_col + 1) / 2
                        center_row = num_rows - (merge_min_row + merge_max_row + 1) / 2
                        
                        ax.text(
                            center_col * cell_width,
                            center_row * cell_height,
                            text,
                            ha='center', va='center',
                            fontsize=9,
                            fontweight='bold',
                            wrap=True,
                            bbox=dict(boxstyle="round,pad=0.1",
                                    facecolor='white',
                                    alpha=0.8)
                        )
        
        # セルデータを表示（ページ範囲内のもののみ）
        for coord, value in sheet_data['cells'].items():
            import re
            match = re.match(r'([A-Z]+)(\d+)', coord)
            if match:
                col = column_index_from_string(match.group(1))
                row = int(match.group(2))
                
                # ページ範囲内かチェック
                if (page_min_row <= row <= page_max_row and
                    page_min_col <= col <= page_max_col):
                    
                    # ページ座標に変換
                    plot_col = col - page_min_col
                    plot_row = row - page_min_row
                    
                    # 値を文字列に変換
                    value_str = str(value)
                    
                    # 長いテキストは短縮
                    max_chars = int(cell_width * 6)
                    if len(value_str) > max_chars:
                        value_str = value_str[:max_chars-2] + '..'
                    
                    # フォントサイズを調整
                    if len(value_str) < 10:
                        fontsize = 9
                    elif len(value_str) < 20:
                        fontsize = 8
                    else:
                        fontsize = 7
                    
                    # セルに値を表示
                    ax.text(
                        (plot_col + 0.5) * cell_width,
                        (num_rows - plot_row - 0.5) * cell_height,
                        value_str,
                        ha='center', va='center',
                        fontsize=fontsize,
                        wrap=False,
                        bbox=dict(boxstyle="round,pad=0.05",
                                facecolor='white',
                                alpha=0.6)
                    )
        
        # 軸の設定
        ax.set_aspect('auto')
        
        # 列ラベル
        col_labels = [get_column_letter(i) for i in range(page_min_col, page_max_col + 1)]
        ax.set_xticks(np.arange(0.5, num_cols, 1) * cell_width)
        ax.set_xticklabels(col_labels, fontsize=9)
        ax.set_xlabel('列', fontsize=11)
        
        # 行ラベル
        row_labels = list(range(page_max_row, page_min_row - 1, -1))
        ax.set_yticks(np.arange(0.5, num_rows, 1) * cell_height)
        ax.set_yticklabels(row_labels, fontsize=9)
        ax.set_ylabel('行', fontsize=11)
        
        # タイトル（ページ情報を含む）
        title = f'{sheet.title} - ページ {page_num}/{total_pages}'
        title += f' (範囲: {get_column_letter(page_min_col)}{page_min_row}:{get_column_letter(page_max_col)}{page_max_row})'
        ax.set_title(title, fontsize=12, fontweight='bold', pad=10)
        
        # ページ情報をフッターに追加
        footer_text = f'ページ {page_num} / {total_pages}'
        ax.text(0.5, -0.08, footer_text, transform=ax.transAxes,
                ha='center', fontsize=10,
                bbox=dict(boxstyle="round,pad=0.3",
                        facecolor='#E3F2FD',
                        alpha=0.8))
        
        plt.tight_layout()
        
        return fig
        """
        AI処理用の強化された可視化画像を作成（内容重視版）
        
        Args:
            sheet_name (str): シート名
            output_path (str): 出力パス
            show_structure (bool): 構造情報を表示
            dpi (int): 画像の解像度
            show_plot (bool): 画像を表示するか（False=保存のみ）
        """
        if sheet_name not in self.ai_optimized_data:
            self.extract_structured_data(sheet_name)
        
        sheet = self.workbook[sheet_name]
        sheet_data = self.ai_optimized_data[sheet_name]
        
        # データがある範囲を検出
        min_data_row = float('inf')
        max_data_row = 0
        min_data_col = float('inf')
        max_data_col = 0
        
        for coord in sheet_data['cells'].keys():
            import re
            match = re.match(r'([A-Z]+)(\d+)', coord)
            if match:
                col = column_index_from_string(match.group(1))
                row = int(match.group(2))
                min_data_row = min(min_data_row, row)
                max_data_row = max(max_data_row, row)
                min_data_col = min(min_data_col, col)
                max_data_col = max(max_data_col, col)
        
        # 表示範囲を実データに合わせる（余白を追加）
        margin = 2
        display_min_row = max(1, min_data_row - margin)
        display_max_row = min(sheet.max_row, max_data_row + margin)
        display_min_col = max(1, min_data_col - margin)
        display_max_col = min(sheet.max_column, max_data_col + margin)
        
        # 表示する行数と列数
        num_rows = display_max_row - display_min_row + 1
        num_cols = display_max_col - display_min_col + 1
        
        # セルサイズを内容に応じて動的に調整
        base_cell_width = 2.5
        base_cell_height = 0.8
        
        # 最大文字長を確認してセル幅を調整
        max_text_length = 10
        for coord, value in sheet_data['cells'].items():
            max_text_length = max(max_text_length, len(str(value)))
        
        cell_width = base_cell_width * min(2.0, max_text_length / 10)
        cell_height = base_cell_height
        
        # 図のサイズを計算
        fig_width = max(12, min(num_cols * cell_width, 35))
        fig_height = max(8, min(num_rows * cell_height, 25))
        
        fig, ax = plt.subplots(figsize=(fig_width, fig_height))
        
        # 座標系の設定（表示範囲に合わせる）
        ax.set_xlim(0, num_cols * cell_width)
        ax.set_ylim(0, num_rows * cell_height)
        
        # 背景グリッド
        for i in range(num_rows + 1):
            ax.axhline(y=i * cell_height, color='#D0D0D0', linewidth=0.8)
        for j in range(num_cols + 1):
            ax.axvline(x=j * cell_width, color='#D0D0D0', linewidth=0.8)
        
        # テーブル領域を描画
        colors = ['#E8F4FD', '#E8F5E9', '#FFF8E1', '#FCE4EC', '#F3E5F5']
        
        # 内部詳細データまたは公開データから境界を取得
        if '_internal' in sheet_data and 'tables_detail' in sheet_data['_internal']:
            tables_detail = sheet_data['_internal']['tables_detail']
        else:
            tables_detail = []
            for table_info in sheet_data['tables']:
                range_parts = table_info['range'].split(':')
                if len(range_parts) == 2:
                    import re
                    start_match = re.match(r'([A-Z]+)(\d+)', range_parts[0])
                    end_match = re.match(r'([A-Z]+)(\d+)', range_parts[1])
                    if start_match and end_match:
                        tables_detail.append({
                            'bounds': {
                                'min_row': int(start_match.group(2)),
                                'max_row': int(end_match.group(2)),
                                'min_col': column_index_from_string(start_match.group(1)),
                                'max_col': column_index_from_string(end_match.group(1))
                            }
                        })
        
        # テーブル領域を描画
        for idx, table in enumerate(tables_detail):
            if 'bounds' in table:
                bounds = table['bounds']
                # 表示範囲内のテーブルのみ描画
                if (bounds['min_row'] <= display_max_row and 
                    bounds['max_row'] >= display_min_row and
                    bounds['min_col'] <= display_max_col and
                    bounds['max_col'] >= display_min_col):
                    
                    # 表示座標に変換
                    table_min_row = max(bounds['min_row'], display_min_row) - display_min_row
                    table_max_row = min(bounds['max_row'], display_max_row) - display_min_row
                    table_min_col = max(bounds['min_col'], display_min_col) - display_min_col
                    table_max_col = min(bounds['max_col'], display_max_col) - display_min_col
                    
                    rect = Rectangle(
                        (table_min_col * cell_width, 
                         (num_rows - table_max_row - 1) * cell_height),
                        (table_max_col - table_min_col + 1) * cell_width,
                        (table_max_row - table_min_row + 1) * cell_height,
                        facecolor=colors[idx % len(colors)],
                        edgecolor='#2196F3',
                        linewidth=2.5,
                        alpha=0.4
                    )
                    ax.add_patch(rect)
                    
                    # テーブルラベル
                    ax.text(
                        (table_min_col - 0.3) * cell_width,
                        (num_rows - table_min_row + 0.3) * cell_height,
                        f"Table {idx+1}",
                        fontsize=11,
                        fontweight='bold',
                        color='#1565C0',
                        bbox=dict(boxstyle="round,pad=0.3", 
                                facecolor='white', 
                                edgecolor='#1565C0',
                                alpha=0.9)
                    )
        
        # 結合セルを描画
        for merged in sheet_data['merged']:
            import re
            range_match = re.match(r'([A-Z]+)(\d+):([A-Z]+)(\d+)', merged['range'])
            if range_match:
                m_min_col = column_index_from_string(range_match.group(1))
                m_min_row = int(range_match.group(2))
                m_max_col = column_index_from_string(range_match.group(3))
                m_max_row = int(range_match.group(4))
                
                # 表示範囲内の場合のみ描画
                if (m_min_row <= display_max_row and 
                    m_max_row >= display_min_row and
                    m_min_col <= display_max_col and
                    m_max_col >= display_min_col):
                    
                    # 表示座標に変換
                    merge_min_row = max(m_min_row, display_min_row) - display_min_row
                    merge_max_row = min(m_max_row, display_max_row) - display_min_row
                    merge_min_col = max(m_min_col, display_min_col) - display_min_col
                    merge_max_col = min(m_max_col, display_max_col) - display_min_col
                    
                    rect = Rectangle(
                        (merge_min_col * cell_width,
                         (num_rows - merge_max_row - 1) * cell_height),
                        (merge_max_col - merge_min_col + 1) * cell_width,
                        (merge_max_row - merge_min_row + 1) * cell_height,
                        facecolor='#FFF59D',
                        edgecolor='#F57C00',
                        linewidth=2,
                        alpha=0.5
                    )
                    ax.add_patch(rect)
                    
                    # 結合セルの値を表示
                    if merged['value']:
                        text = str(merged['value'])
                        center_col = (merge_min_col + merge_max_col + 1) / 2
                        center_row = num_rows - (merge_min_row + merge_max_row + 1) / 2
                        
                        ax.text(
                            center_col * cell_width,
                            center_row * cell_height,
                            text,
                            ha='center', va='center',
                            fontsize=10,
                            fontweight='bold',
                            wrap=True,
                            bbox=dict(boxstyle="round,pad=0.2", 
                                    facecolor='white', 
                                    alpha=0.8)
                        )
        
        # すべてのセルデータを表示（表示範囲内のもの）
        for coord, value in sheet_data['cells'].items():
            import re
            match = re.match(r'([A-Z]+)(\d+)', coord)
            if match:
                col = column_index_from_string(match.group(1))
                row = int(match.group(2))
                
                # 表示範囲内かチェック
                if (display_min_row <= row <= display_max_row and 
                    display_min_col <= col <= display_max_col):
                    
                    # 表示座標に変換
                    plot_col = col - display_min_col
                    plot_row = row - display_min_row
                    
                    # 値を文字列に変換
                    value_str = str(value)
                    
                    # 長いテキストは折り返し
                    max_chars = int(cell_width * 5)
                    if len(value_str) > max_chars:
                        # 複数行に分割
                        lines = []
                        for i in range(0, len(value_str), max_chars):
                            lines.append(value_str[i:i+max_chars])
                        value_str = '\n'.join(lines[:3])  # 最大3行
                        if len(lines) > 3:
                            value_str += '...'
                    
                    # フォントサイズを内容に応じて調整
                    if len(value_str) < 10:
                        fontsize = 9
                    elif len(value_str) < 20:
                        fontsize = 8
                    else:
                        fontsize = 7
                    
                    # セルに値を表示
                    ax.text(
                        (plot_col + 0.5) * cell_width,
                        (num_rows - plot_row - 0.5) * cell_height,
                        value_str,
                        ha='center', va='center',
                        fontsize=fontsize,
                        wrap=True,
                        bbox=dict(boxstyle="round,pad=0.05", 
                                facecolor='white', 
                                alpha=0.7)
                    )
        
        # 軸の設定
        ax.set_aspect('auto')
        
        # 列ラベル（表示範囲に合わせる）
        col_labels = [get_column_letter(i) for i in range(display_min_col, display_max_col + 1)]
        ax.set_xticks(np.arange(0.5, num_cols, 1) * cell_width)
        ax.set_xticklabels(col_labels, fontsize=9)
        ax.set_xlabel('列', fontsize=11)
        
        # 行ラベル（表示範囲に合わせる）
        row_labels = list(range(display_max_row, display_min_row - 1, -1))
        ax.set_yticks(np.arange(0.5, num_rows, 1) * cell_height)
        ax.set_yticklabels(row_labels, fontsize=9)
        ax.set_ylabel('行', fontsize=11)
        
        # タイトル
        title = f'シート: {sheet_name}'
        if sheet_data['cells']:
            title += f' (データ範囲: {get_column_letter(min_data_col)}{min_data_row}:{get_column_letter(max_data_col)}{max_data_row})'
        ax.set_title(title, fontsize=14, fontweight='bold', pad=15)
        
        # 統計情報
        stats_text = f"セル数: {len(sheet_data['cells'])} | テーブル: {len(sheet_data['tables'])} | 結合セル: {len(sheet_data['merged'])}"
        ax.text(0.5, -0.05, stats_text, transform=ax.transAxes,
                ha='center', fontsize=10,
                bbox=dict(boxstyle="round,pad=0.3", 
                        facecolor='#E3F2FD', 
                        alpha=0.8))
        
        plt.tight_layout()
        
        if output_path:
            plt.savefig(output_path, dpi=dpi, bbox_inches='tight', 
                       facecolor='white', edgecolor='none')
            print(f"画像を保存しました: {output_path}")
        
        if show_plot:
            plt.show()
        else:
            plt.close(fig)
        
        return fig
    
    def export_for_ai_processing(self, output_dir="ai_input", show_images=False, 
                                compact_json=True, use_pagination=True,
                                rows_per_page=30, cols_per_page=10,
                                use_print_area=True):
        """
        AI処理用にデータをエクスポート（ページ単位のデータ紐付け対応）
        
        Args:
            output_dir (str): 出力ディレクトリ
            show_images (bool): 画像を表示するか
            compact_json (bool): JSONを最小化するか
            use_pagination (bool): ページ分割を使用するか
            rows_per_page (int): 1ページあたりの行数
            cols_per_page (int): 1ページあたりの列数
            use_print_area (bool): 印刷範囲のみを対象とするか
        """
        output_path = Path(output_dir)
        output_path.mkdir(exist_ok=True)
        
        # 全シートのデータを抽出（印刷範囲対応）
        self.extract_structured_data(use_print_area=use_print_area)
        
        # エクスポート用データ
        export_data = {}
        
        for sheet_name in self.sheet_names:
            if sheet_name not in self.ai_optimized_data:
                print(f"警告: シート '{sheet_name}' のデータが見つかりません。スキップします。")
                continue
            sheet_data = self.ai_optimized_data[sheet_name]
            
            if use_pagination:
                # ページ分割してデータと画像を生成
                print(f"\n{sheet_name} をページ分割中...")
                pages_data = self.create_paged_visualizations_with_data(
                    sheet_name, 
                    output_dir=output_dir,
                    rows_per_page=rows_per_page,
                    cols_per_page=cols_per_page,
                    dpi=120
                )
                
                # ページ単位のデータ構造
                export_data[sheet_name] = {
                    'meta': {
                        'print_area': sheet_data['print_area'],
                        'data_bounds': sheet_data['data_bounds'],
                        'total_cells': len(sheet_data['cells']),
                        'total_tables': len(sheet_data['tables']),
                        'total_merged': len(sheet_data['merged']),
                        'page_count': len(pages_data)
                    },
                    'pages': pages_data
                }
            else:
                # 従来の単一構造
                export_data[sheet_name] = {
                    'cells': sheet_data['cells'],
                    'tables': sheet_data['tables'],
                    'merged': sheet_data['merged'],
                    'print_area': sheet_data['print_area'],
                    'data_bounds': sheet_data['data_bounds']
                }
                
                # 単一画像生成
                img_path = output_path / f"{sheet_name}_structure.png"
                self.create_enhanced_visualization(
                    sheet_name, 
                    str(img_path),
                    show_plot=show_images
                )
                export_data[sheet_name]['visualization'] = f"{sheet_name}_structure.png"
        
        # メインJSONファイルとして保存
        if use_pagination:
            # ページ単位のデータ構造で保存
            json_path = output_path / "paged_data.json"
            with open(json_path, 'w', encoding='utf-8') as f:
                if compact_json:
                    json.dump(export_data, f, ensure_ascii=False, separators=(',', ':'), default=str)
                else:
                    json.dump(export_data, f, ensure_ascii=False, indent=2, default=str)
            
            # 各ページのデータを個別ファイルでも保存（オプション）
            if not compact_json:
                for sheet_name, sheet_data in export_data.items():
                    if 'pages' in sheet_data:
                        sheet_dir = output_path / sheet_name
                        sheet_dir.mkdir(exist_ok=True)
                        for page_key, page_data in sheet_data['pages'].items():
                            page_json = sheet_dir / f"{page_key}.json"
                            with open(page_json, 'w', encoding='utf-8') as f:
                                json.dump(page_data, f, ensure_ascii=False, indent=2, default=str)
        else:
            # 従来の構造で保存
            json_path = output_path / "structured_data.json"
            with open(json_path, 'w', encoding='utf-8') as f:
                if compact_json:
                    json.dump(export_data, f, ensure_ascii=False, separators=(',', ':'), default=str)
                else:
                    json.dump(export_data, f, ensure_ascii=False, indent=2, default=str)
        
        # Markdown変換ガイドを生成
        guide_path = output_path / "conversion_guide.md"
        with open(guide_path, 'w', encoding='utf-8') as f:
            f.write(self._generate_paged_guide(export_data, use_pagination))
        
        print(f"\n✅ AI処理用データを出力しました:")
        print(f"  - JSONデータ: {json_path}")
        
        # ファイルサイズを表示
        file_size = json_path.stat().st_size
        if file_size < 1024:
            print(f"    サイズ: {file_size} bytes")
        elif file_size < 1024 * 1024:
            print(f"    サイズ: {file_size / 1024:.1f} KB")
        else:
            print(f"    サイズ: {file_size / (1024 * 1024):.1f} MB")
        
        if use_pagination:
            total_pages = sum(len(s['pages']) for s in export_data.values() if 'pages' in s)
            print(f"  - 画像ファイル: {total_pages}ページ分を生成")
            print(f"  - データ構造: ページ単位でデータを紐付け")
        else:
            print(f"  - 画像ファイル: {output_path}/*.png")
        
        print(f"  - 変換ガイド: {guide_path}")
        
        if use_print_area:
            print(f"  - 対象範囲: 印刷範囲内のみ")
        
        return export_data
    
    def _generate_paged_guide(self, data, use_pagination):
        """
        ページ単位データ構造のMarkdown変換ガイドを生成
        
        Args:
            data: エクスポートデータ
            use_pagination: ページ分割を使用したか
            
        Returns:
            str: Markdownガイド
        """
        guide = """# Excel → Markdown 変換ガイド

## データ構造
"""
        
        if use_pagination:
            guide += """### ページ単位のデータ構造
```json
{
  "Sheet1": {
    "meta": {
      "print_area": "A1:Z100",        // 印刷範囲
      "data_bounds": {...},            // 実データ範囲
      "total_cells": 500,              // 総セル数
      "page_count": 6                  // ページ数
    },
    "pages": {
      "page_001": {
        "image": "Sheet1_page_001.png",
        "range": "A1:J30",             // このページの範囲
        "cells": {"A1": "値", ...},    // このページのセルデータ
        "tables": [...],                // このページに含まれるテーブル
        "merged": [...],                // このページの結合セル
        "stats": {...}                  // 統計情報
      },
      "page_002": {...}
    }
  }
}
```

### 特徴
- **印刷範囲対応**: Excelの印刷範囲内のデータのみを抽出
- **ページ単位のデータ**: 各ページの画像とデータが1対1で対応
- **オーバーラップ**: ページ間に若干の重複があり、連続性を保持

## 変換手順
1. 各ページの画像とデータを確認
2. ページごとに含まれるテーブルをMarkdown形式に変換
3. 結合セルを見出しやタイトルとして配置
4. ページ間の重複部分は一度だけ出力
5. 全ページを統合して完成

### ページ処理の例（Python）
```python
import json

with open('paged_data.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

for sheet_name, sheet_data in data.items():
    print(f"# {sheet_name}")
    
    for page_key, page_data in sheet_data['pages'].items():
        print(f"\\n## {page_data['range']}")
        
        # このページのセルを処理
        for coord, value in page_data['cells'].items():
            print(f"{coord}: {value}")
        
        # このページのテーブルを処理
        for table in page_data['tables']:
            print(f"Table {table['id']}: {table['range']}")
```
"""
        else:
            guide += """### 従来のデータ構造
```json
{
  "Sheet1": {
    "cells": {"A1": "値", "B2": "値2", ...},
    "tables": [{"id": "T1", "range": "A5:D10", "header": true}],
    "merged": [{"range": "A1:D1", "value": "タイトル"}],
    "print_area": "A1:Z100",
    "data_bounds": {"range": "A1:M50", ...}
  }
}
```

## 変換手順
1. JSONファイルの`cells`辞書からセルデータを取得
2. `tables`の範囲情報を使ってテーブルを構築
3. `merged`の値を見出しとして配置
"""
        
        guide += """

## AIへの指示例
"""
        
        if use_pagination:
            guide += """```
以下のページ単位のデータと画像からMarkdownを生成してください：

1. paged_data.jsonを読み込む
2. 各ページの画像（Sheet1_page_001.png等）とそのページのデータを対応させて確認
3. ページごとのcellsデータからテーブル構造を復元
4. tablesの範囲情報を使ってMarkdownテーブルを構築
5. mergedの値を適切な見出しレベルで配置
6. ページ間の重複は除去して、連続した文書として出力

注意: 各ページのデータは独立しているため、ページごとに処理してから統合してください。
```"""
        else:
            guide += """```
JSONデータと画像からMarkdownを生成してください：

1. structured_data.jsonのcellsからデータを取得
2. tablesの範囲に基づいてMarkdownテーブルを構築
3. mergedを見出しとして配置
```"""
        
        return guide
    
    def _generate_markdown_hints(self, sheet_data):
        """
        Markdown変換のヒントを生成
        
        Args:
            sheet_data: シートデータ
            
        Returns:
            dict: Markdownヒント
        """
        hints = {
            'suggested_structure': [],
            'table_conversion': [],
            'formatting_rules': []
        }
        
        # ドキュメント構造の提案
        if sheet_data['semantic_structure']['sections']:
            hints['suggested_structure'].append({
                'type': 'hierarchical',
                'sections': len(sheet_data['semantic_structure']['sections']),
                'use_headers': True
            })
        
        # テーブル変換の提案
        for table in sheet_data['tables']:
            if table['has_header']:
                hints['table_conversion'].append({
                    'table_id': table.get('table_id'),
                    'format': 'markdown_table',
                    'alignment': table['alignment']
                })
            else:
                hints['table_conversion'].append({
                    'table_id': table.get('table_id'),
                    'format': 'list' if table['alignment'] == 'vertical' else 'key_value'
                })
        
        # フォーマットルール
        if sheet_data['formatting_hints']:
            bold_count = sum(1 for h in sheet_data['formatting_hints'] if h.get('bold'))
            if bold_count > 0:
                hints['formatting_rules'].append('Use **bold** for emphasized text')
            
            heading_count = sum(1 for h in sheet_data['formatting_hints'] 
                              if h.get('possible_heading'))
            if heading_count > 0:
                hints['formatting_rules'].append('Convert large/bold text to headers (#, ##, ###)')
        
        return hints
    
    def _generate_ai_prompt_template(self, integrated_data):
        """
        AI用のプロンプトテンプレートを生成
        
        Args:
            integrated_data: 統合データ
            
        Returns:
            str: プロンプトテンプレート
        """
        prompt = f"""# Excel to Markdown Conversion Task

## Source Information
- File: {integrated_data['source_file']}
- Total Sheets: {integrated_data['total_sheets']}

## Conversion Instructions

Please convert the following Excel data to well-structured Markdown format:

### Guidelines:
1. **Preserve Structure**: Maintain the logical structure of the original document
2. **Table Handling**: 
   - Convert tables with headers to Markdown tables
   - Convert key-value pairs to definition lists or formatted text
   - For complex tables, consider breaking them into sections
3. **Formatting**:
   - Use headers (# ## ###) for sections and titles
   - Apply **bold** and *italic* where emphasized in original
   - Preserve lists and numbered items
4. **Merged Cells**: Treat as section headers or important notes
5. **Data Types**: Format numbers, percentages, and dates appropriately

## Sheet-Specific Information:
"""
        
        for sheet_info in integrated_data['sheets']:
            sheet_data = sheet_info['structure']
            prompt += f"""
### Sheet: {sheet_info['name']}
- Document Type: {sheet_data['semantic_structure']['document_type']}
- Tables: {len(sheet_data['tables'])}
- Sections: {len(sheet_data['semantic_structure']['sections'])}
- Suggested Approach: {sheet_info['markdown_hints']['suggested_structure']}

"""
        
        prompt += """
## Expected Output Format:

```markdown
# [Document Title]

## [Section 1]
[Content...]

### [Subsection]
[Table or formatted data]

## [Section 2]
[Content...]
```

Please analyze the provided JSON structure and images to create an accurate Markdown representation.
"""
        
        return prompt


# 使用例
def main():
    """メイン処理"""
    
    # Excelファイルのパスを指定
    excel_file = "sample.xlsx"  # 実際のファイルパスに変更してください
    
    try:
        # インスタンスの作成
        processor = ExcelToMarkdownPreprocessor(excel_file)
        
        # ページ分割でAI処理用データをエクスポート
        print("ページ分割してAI処理用データを生成中...")
        integrated_data = processor.export_for_ai_processing(
            "ai_input",
            show_images=False,      # 画像は表示しない
            compact_json=True,      # JSONを最小化
            use_pagination=True,    # ページ分割を使用
            rows_per_page=30,       # 1ページあたり30行
            cols_per_page=10        # 1ページあたり10列
        )
        
        # 各シートの簡易サマリー
        print("\n=== 抽出結果 ===")
        for sheet_name, sheet_data in integrated_data.items():
            print(f"\n【{sheet_name}】")
            if isinstance(sheet_data, dict):
                # cellsキーが存在するかチェック
                if 'cells' in sheet_data:
                    print(f"  セル数: {len(sheet_data['cells'])}")
                else:
                    print(f"  セル数: データなし")
                    
                # tablesキーが存在するかチェック  
                if 'tables' in sheet_data:
                    print(f"  テーブル数: {len(sheet_data['tables'])}")
                    if sheet_data['tables']:
                        for table in sheet_data['tables']:
                            if isinstance(table, dict) and 'id' in table and 'range' in table:
                                print(f"    - {table['id']}: {table['range']}")
                else:
                    print(f"  テーブル数: データなし")
                    
                # mergedキーが存在するかチェック
                if 'merged' in sheet_data:
                    print(f"  結合セル数: {len(sheet_data['merged'])}")
                else:
                    print(f"  結合セル数: データなし")
            else:
                print(f"  データ形式が予期されるものと異なります: {type(sheet_data)}")
                continue
            
            # ページ情報
            if 'pages' in sheet_data:
                pages = sheet_data['pages']
                print(f"  生成ページ数: {len(pages)}")
                
                if isinstance(pages, list):
                    # リスト形式の場合
                    if len(pages) <= 5:
                        for page in pages:
                            print(f"    - {page}")
                    else:
                        print(f"    - {pages[0]} ... {pages[-1]}")
                elif isinstance(pages, dict):
                    # 辞書形式の場合
                    page_keys = list(pages.keys())
                    if len(page_keys) <= 5:
                        for key in page_keys:
                            print(f"    - {key}: {pages[key]}")
                    else:
                        print(f"    - {page_keys[0]} ... {page_keys[-1]}")
                else:
                    print(f"    - ページデータ形式不明: {type(pages)}")
            
            # サンプルデータ表示（最初の3個）
            if 'cells' in sheet_data and sheet_data['cells']:
                print(f"  サンプルデータ:")
                for i, (coord, value) in enumerate(list(sheet_data['cells'].items())[:3]):
                    value_str = str(value)[:30] + '...' if len(str(value)) > 30 else str(value)
                    print(f"    {coord}: {value_str}")
        
        print("\n✅ 処理完了！")
        print("生成されたJSONファイルとページ画像をAIに入力してMarkdown変換を依頼してください。")
        
        # 単一画像版も必要な場合のオプション
        print("\n📌 単一画像版が必要な場合は以下を実行:")
        print("processor.export_for_ai_processing('ai_input_single', use_pagination=False)")
        
    except FileNotFoundError:
        print(f"エラー: ファイル '{excel_file}' が見つかりません。")
    except Exception as e:
        print(f"エラーが発生しました: {e}")
        import traceback
        traceback.print_exc()


        print(f"エラー: ファイル '{excel_file}' が見つかりません。")
    except Exception as e:
        print(f"エラーが発生しました: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
